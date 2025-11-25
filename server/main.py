from __future__ import annotations
from io import BytesIO
import secrets
import traceback
from typing import Dict, List

import pandas as pd
from fastapi import FastAPI, File, UploadFile, Request, Query
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
import math

from wb_io.wb_readers import (
    read_stock_history,
    read_stock_snapshot,
    read_intransit_file,
    read_fulfillment_stock_file,
)
from engine.transform import (
    sales_by_warehouse_from_details,
    merge_sales_with_stock_today,
    stock_from_snapshot,
)
from engine.export_prototype import build_prototype_workbook

# Название листа с продажами в итоговом Excel (для логов)
SHEET_SALES_NAME = "Продажи по складам"

_SALES_STOCK_COLUMNS = [
    "Артикул продавца",
    "Артикул WB",
    "Склад",
    "Заказали, шт",
    "Остаток на сегодня",
]

app = FastAPI()
templates = Jinja2Templates(directory="server/templates")

try:
    app.mount("/static", StaticFiles(directory="server/static"), name="static")
except Exception:
    pass

_memory_artifacts: Dict[str, bytes] = {}

# (FF‑хелперы удалены как неиспользуемые)

def _distribute_ff_for_sku(df, ff_total: float):
    """
    Распределяет ограниченный остаток FF по строкам одного SKU.
    Ожидаемые колонки в df:
      - 'Рекомендация, шт'  (теоретический заказ по складу)
      - 'Вес склада'        (нормированный вес среди выбранных складов)
      - 'MOQ'               (кратность поставки, шаг округления)

    Возвращает pd.Series той же длины, что df, с количеством
    для колонки 'Рекомендация с учётом ФФ'.
    """
    import pandas as pd  # защищаемся на случай переименований

    if df is None or df.empty:
        return pd.Series([0] * 0, index=df.index if df is not None else None)

    base = pd.to_numeric(df.get("Рекомендация, шт"), errors="coerce").fillna(0.0)
    base = base.astype(float).clip(lower=0.0)
    if ff_total is None or ff_total <= 0:
        return pd.Series(0, index=df.index, dtype="int64")

    total_base = float(base.sum())
    if total_base <= 0:
        return pd.Series(0, index=df.index, dtype="int64")

    # Если ФФ хватает на весь теоретический спрос — возвращаем базовый заказ как есть
    if ff_total >= total_base:
        return base.round().astype("int64")

    weights = pd.to_numeric(df.get("Вес склада"), errors="coerce").fillna(0.0)
    weights = weights.astype(float).clip(lower=0.0)
    total_w = float(weights.sum())
    if total_w > 0:
        shares = weights / total_w
    else:
        # fallback: распределяем пропорционально теоретическому спросу
        shares = base / total_base

    raw = ff_total * shares
    # не выдаём больше, чем теоретический заказ
    raw = raw.where(raw <= base, base)

    moq = pd.to_numeric(df.get("MOQ"), errors="coerce").fillna(0.0)
    rounded = []
    for qty, step in zip(raw, moq):
        q = float(qty)
        s = float(step) if step and step > 0 else 1.0
        if q <= 0:
            rounded.append(0.0)
        else:
            rounded.append(math.ceil(q / s) * s)

    return pd.Series(rounded, index=df.index).fillna(0.0).astype("int64")


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/download/fulfillment_template.xlsx")
async def download_fulfillment_template():
    """Генерирует и отдает XLSX-шаблон «Остатки Фулфилмент»"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Шаблон"
    worksheet.append(["Артикул продавца", "Артикул WB", "Количество"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="fulfillment_template.xlsx"'
        },
    )

# [WB_ANCHOR] build endpoint
@app.post("/build")
async def build(
    files: List[UploadFile] = File(...),
    dummy: int = Query(0, ge=0, le=1),
):
    logs: List[str] = []
    combined_frames: List[pd.DataFrame] = []
    supplies_frames: List[pd.DataFrame] = []
    fulfillment_frames: List[pd.DataFrame] = []
    daily_frames: List[pd.DataFrame] = []  # История остатков по дням (по сети)
    sku_ref: pd.DataFrame | None = None
    sku_map: Dict[str, Dict[str, str | None]] = {}
    unknown_sku: List[str] = []
    unknown_seen: set[tuple[str, str | None, str | None, str | None]] = set()

    if not files and dummy != 1:
        return JSONResponse({"ok": False, "log": "Файлы не переданы"}, status_code=400)

    try:
        if dummy == 1:
            logs.append("SMOKE: dummy=1 — сборка на встроенных фикстурах")
            daily_frames.append(
                pd.DataFrame(
                    {
                        "Артикул продавца": ["DUMMY_SKU"],
                        "Артикул WB": ["000000"],
                        "01.10.2025": [10],
                        "02.10.2025": [12],
                        "03.10.2025": [0],
                    }
                )
            )
            detail_dummy = pd.DataFrame(
                {
                    "Склад": ["Тула", "Казань"],
                    "Артикул продавца": ["DUMMY_SKU", "DUMMY_SKU"],
                    "Артикул WB": ["000000", "000000"],
                    "Заказали, шт": [8, 4],
                }
            )
            merged = merge_sales_with_stock_today(detail_dummy, detail_dummy, daily_frames[0])
            if not merged.empty:
                combined_frames.append(merged)
            logs.append("SMOKE: добавлены фикстуры — продажи и остатки по дням (1 SKU)")
        else:
            for f in files:
                raw = await f.read()

                # 0) Справочник SKU — удалено: справочник отключён

                # 1) СНАЧАЛА: «Остатки Фулфилмент» (простой файл с двумя колонками)
                fulfillment_one = read_fulfillment_stock_file(raw, f.filename)
                if not fulfillment_one.empty:
                    fulfillment_frames.append(fulfillment_one)
                    logs.append(
                        f"{f.filename}: источник «Остатки Фулфилмент» — {len(fulfillment_one)} строк"
                    )
                    continue

                # 2) ЗАТЕМ: «Поставки в пути»
                intransit_df = read_intransit_file(raw, f.filename)
                if intransit_df.attrs.get("intransit"):
                    supplies_frames.append(intransit_df)
                    logs.append(
                        f"{f.filename}: источник «Поставки в пути» — {len(intransit_df)} строк"
                    )
                    continue

                # 3) ПОТОМ: «Остатки по складам» (снимок)
                snapshot_df = read_stock_snapshot(raw, f.filename)
                if snapshot_df is not None:
                    df_stock = stock_from_snapshot(snapshot_df)
                    if not df_stock.empty:
                        df_stock = df_stock.rename(columns={"Остаток": "Остаток на сегодня"})
                        df_stock["Заказали, шт"] = 0
                        df_stock["Склад"] = df_stock["Склад"].fillna("-").replace("", "-")
                        df_stock = df_stock[_SALES_STOCK_COLUMNS]
                        combined_frames.append(df_stock)
                    logs.append(
                        f"{f.filename}: источник «Остатки по складам» — {len(df_stock)} строк"
                    )
                    continue

                # 4) ИНАЧЕ: «История остатков» (детали + по дням)
                sheets = read_stock_history(raw, f.filename)
                if not sheets:
                    logs.append(f"{f.filename}: источник не распознан")
                    continue

                detail = sheets.get("Детальная информация")
                daily = sheets.get("Остатки по дням")

                df_sales = (
                    sales_by_warehouse_from_details(detail)
                    if detail is not None
                    else pd.DataFrame()
                )
                merged = merge_sales_with_stock_today(df_sales, detail, daily)
                if not merged.empty:
                    combined_frames.append(merged)

                logs.append(
                    f"{f.filename}: источник «История остатков» — {len(merged)} строк"
                )
                # Сбор дневных остатков (по сети) для вкладки «История остатков по дням»
                if daily is not None and not daily.empty:
                    df = daily.copy()
                    id_cols = ["Артикул продавца", "Артикул WB"]
                    date_cols = [c for c in df.columns if c not in id_cols]
                    if date_cols:
                        for c in date_cols:
                            df[c] = pd.to_numeric(
                                df[c]
                                .astype(str)
                                .str.replace("\xa0", "", regex=False)
                                .str.replace(" ", "", regex=False)
                                .str.replace(",", ".", regex=False),
                                errors="coerce",
                            ).fillna(0)
                        daily_frames.append(df[id_cols + date_cols])

        # (удалено) SKU‑валидация отключена: сборка не зависит от справочника.
        if combined_frames:
            sales_stock = pd.concat(combined_frames, ignore_index=True).reindex(columns=_SALES_STOCK_COLUMNS)
        else:
            sales_stock = pd.DataFrame(columns=_SALES_STOCK_COLUMNS)

        logs.append(f"Итог: «{SHEET_SALES_NAME}» — {len(sales_stock)}.")

        # Свести «Историю остатков по дням» из всех загруженных файлов
        daily_stock_df = None
        if daily_frames:
            base = daily_frames[0]
            for extra in daily_frames[1:]:
                base = base.merge(
                    extra,
                    on=["Артикул продавца", "Артикул WB"],
                    how="outer",
                )
            id_cols = ["Артикул продавца", "Артикул WB"]
            non_id = [c for c in base.columns if c not in id_cols]
            if non_id:
                daily_stock_df = (
                    base.groupby(id_cols, dropna=False, as_index=False)[non_id].max()
                )
            else:
                daily_stock_df = base[id_cols].drop_duplicates().reset_index(drop=True)
            logs.append(
                f"Итог: «История остатков по дням» — {len(daily_stock_df)} SKU"
            )

        supplies_df = pd.concat(supplies_frames, ignore_index=True) if supplies_frames else None
        fulfillment_df = (
            pd.concat(fulfillment_frames, ignore_index=True)
            if fulfillment_frames
            else None
        )
        bio: BytesIO = build_prototype_workbook(
            sales_stock,
            supplies_df=supplies_df,
            fulfillment_df=fulfillment_df,
            daily_stock_df=daily_stock_df,
        )
        token = secrets.token_urlsafe(16)
        _memory_artifacts[token] = bio.getvalue()

        if supplies_df is not None:
            logs.append(f"Итог: «Поставки в пути» — {len(supplies_df)} строк")
        if fulfillment_df is not None:
            logs.append(
                f"Итог: «Остатки Фулфилмент» — {len(fulfillment_df)} строк"
            )
        return {"ok": True, "log": "\n".join(logs), "download_token": token}

    except Exception as e:
        tb = traceback.format_exc()
        return JSONResponse(
            {"ok": False, "log": "\n".join(logs + [f'Ошибка: {e}', "TRACEBACK:", tb])},
            status_code=500,
        )

@app.get("/download/{token}")
async def download(token: str):
    blob = _memory_artifacts.pop(token, None)
    if not blob:
        return JSONResponse({"ok": False, "log": "Файл не найден или срок истёк"}, status_code=404)
    # Если формат старый (просто bytes) — используем имя по умолчанию.
    if isinstance(blob, bytes):
        data = blob
        filename = "Input_Prototype_Filled.xlsx"
    else:
        data = blob.get("data")
        filename = blob.get("filename", "result.xlsx")

    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# --------------------------- Рекомендации ---------------------------
@app.post("/recommend")
async def recommend(files: List[UploadFile] = File(...)):
    logs: List[str] = []
    if not files:
        return JSONResponse({"ok": False, "log": "Файл не передан"}, status_code=400)
    try:
        raw: bytes = b""
        f0 = files[0]
        try:
            pos = f0.file.tell()
            f0.file.seek(0)
            raw = f0.file.read()
            f0.file.seek(pos)
        except Exception:
            raw = await f0.read()
            try:
                f0.file.seek(0)
            except Exception:
                pass
        bio = BytesIO(raw)
        xls = pd.ExcelFile(bio)
        # нужен лист «Продажи по складам»
        sheet_name = None
        for name in xls.sheet_names:
            if str(name).strip().lower() == "продажи по складам":
                sheet_name = name
                break
        if sheet_name is None:
            return JSONResponse({"ok": False, "log": "Не найден лист «Продажи по складам» в загруженном файле"}, status_code=400)
        df = xls.parse(sheet_name)
        # нормализуем шапку
        cols = {str(c).strip().lower(): c for c in df.columns}

        def pick(cands):
            for k in cands:
                if k in cols:
                    return cols[k]
            return None

        c_seller = pick(["артикул продавца"])
        c_wb = pick(["артикул wb", "артикул wb."])
        c_wh = pick(["склад"])
        c_avg = pick(["средние продажи в день", "средние продажи/день", "средние продажи"])
        c_coef = pick(["коэф. склада", "коэфф. склада", "коэффициент склада", "коэф", "коэфф"])
        needed = [c_seller, c_wb, c_wh, c_avg]
        if any(c is None for c in needed):
            return JSONResponse(
                {"ok": False, "log": "В листе «Продажи по складам» отсутствуют нужные колонки"},
                status_code=400,
            )
        df_base = pd.DataFrame(
            {
                "Артикул продавца": df[c_seller],
                "Артикул WB": df[c_wb],
                "Склад": df[c_wh],
                "Средние продажи в день": pd.to_numeric(df[c_avg], errors="coerce").fillna(0.0),
            }
        )
        if c_coef is not None:
            df_base["Коэф. склада"] = pd.to_numeric(df[c_coef], errors="coerce").fillna(0.0)
        else:
            df_base["Коэф. склада"] = 0.0
        c_stock = pick(["остаток на сегодня", "остаток"])
        if c_stock is not None:
            df_base["Остаток на сегодня"] = (
                pd.to_numeric(df[c_stock], errors="coerce").fillna(0.0)
            )
        else:
            df_base["Остаток на сегодня"] = 0.0

        # Берём MinStock и MOQ из соответствующих листов
        # --- MinStock ---
        minstock_sheet = None
        for name in xls.sheet_names:
            if str(name).strip().lower() == "minstock":
                minstock_sheet = name
                break
        minstock_map: Dict[tuple[str, str], float] = {}
        if minstock_sheet:
            ms = xls.parse(minstock_sheet)
            cols_ms = {str(c).strip().lower(): c for c in ms.columns}
            c_ms_seller = cols_ms.get("артикул продавца")
            c_ms_wb = cols_ms.get("артикул wb")
            c_ms_val = cols_ms.get("значение")
            if c_ms_seller and c_ms_wb and c_ms_val:
                for _, row in ms.iterrows():
                    key = (str(row[c_ms_seller]).strip(), str(row[c_ms_wb]).strip())
                    try:
                        minstock_map[key] = float(row[c_ms_val])
                    except Exception:
                        continue

        # --- MOQ ---
        moq_sheet = None
        for name in xls.sheet_names:
            if str(name).strip().lower() == "moq":
                moq_sheet = name
                break
        moq_map: Dict[tuple[str, str], float] = {}
        if moq_sheet:
            mo = xls.parse(moq_sheet)
            cols_mo = {str(c).strip().lower(): c for c in mo.columns}
            c_mo_seller = cols_mo.get("артикул продавца")
            c_mo_wb = cols_mo.get("артикул wb")
            c_mo_val = cols_mo.get("moq")
            if c_mo_seller and c_mo_wb and c_mo_val:
                for _, row in mo.iterrows():
                    key = (str(row[c_mo_seller]).strip(), str(row[c_mo_wb]).strip())
                    try:
                        moq_map[key] = float(row[c_mo_val])
                    except Exception:
                        continue

        # --- Окна приёмки ---
        acceptance_days = 0
        acc_sheet = None
        for name in xls.sheet_names:
            if str(name).strip().lower() == "окна приёмки":
                acc_sheet = name
                break
        if acc_sheet:
            ac = xls.parse(acc_sheet)
            cols_ac = {str(c).strip().lower(): c for c in ac.columns}
            c_days = cols_ac.get("количество дней")
            if c_days is not None and not ac.empty:
                try:
                    acceptance_days = int(ac[c_days].iloc[0])
                except Exception:
                    acceptance_days = 0

        # --- Фильтр складов + частота подсортировок ---
        selected_wh: List[str] | None = None
        freq_map: Dict[str, int] = {}
        wh_sheet = None
        for name in xls.sheet_names:
            if str(name).strip().lower() == "склады для подсортировки":
                wh_sheet = name
                break
        if wh_sheet:
            wh = xls.parse(wh_sheet)
            cols_wh = {str(c).strip().lower(): c for c in wh.columns}
            c_wh_name = cols_wh.get("склад")
            c_sel = cols_wh.get("выбрать")
            c_freq = cols_wh.get("частота подсортировок, дни")

            def _is_selected(v: object) -> bool:
                return str(v).strip().lower() in ("1", "true", "да", "истина", "yes")

            if c_wh_name and c_sel:
                selected_wh = []
                for _, row in wh.iterrows():
                    wh_name = str(row[c_wh_name]).strip()
                    if not wh_name:
                        continue
                    if _is_selected(row[c_sel]):
                        selected_wh.append(wh_name)
                    if c_freq:
                        try:
                            freq_map[wh_name] = int(row[c_freq])
                        except Exception:
                            freq_map[wh_name] = freq_map.get(wh_name, 0)
                selected_wh = list(dict.fromkeys(selected_wh))

        if selected_wh:
            logs.append(
                f"Фильтр складов: выбрано {len(selected_wh)} — {', '.join(selected_wh)}"
            )
        else:
            logs.append("Фильтр складов: не задан (используются все склады)")

        # --- Остатки Фулфилмент ---
        ff_stock: Dict[tuple[str, str], float] = {}

        def _norm_id(v):
            if v is None:
                return None
            if isinstance(v, float) and math.isnan(v):
                return None
            s = str(v).strip().replace("\xa0", "")
            s = s.replace(" ", "")
            if s.endswith(".0") and s[:-2].isdigit():
                s = s[:-2]
            return s or None

        try:
            # читаем байты книги надёжно (seek -> read -> seek back)
            pos0 = await files[0].seek(0)
            raw = await files[0].read()
            await files[0].seek(pos0 or 0)
            xl = pd.ExcelFile(BytesIO(raw))
            # ищем лист «Остатки Фулфилмент» без учёта регистра/вариаций
            ff_sheet = None
            for s in xl.sheet_names:
                sl = str(s).strip().lower()
                if sl == "остатки фулфилмент":
                    ff_sheet = s
                    break
            if ff_sheet is None:
                for s in xl.sheet_names:
                    sl = str(s).strip().lower()
                    if "остатк" in sl and "фулф" in sl:
                        ff_sheet = s
                        break
            ff_df = pd.DataFrame()
            if ff_sheet is not None:
                ff_df = xl.parse(ff_sheet)
                ff_df = ff_df.copy()
                ff_df.columns = [str(c).strip() for c in ff_df.columns]
                lc = {str(c).strip().lower(): c for c in ff_df.columns}
                seller_col = lc.get("артикул продавца") or lc.get("артикул поставщика") or lc.get("артикул")
                wb_col = (
                    lc.get("артикул wb")
                    or lc.get("артикул wb.")
                    or lc.get("артикул вб")
                    or lc.get("код товара")
                )
                qty_col = lc.get("количество") or lc.get("остаток") or lc.get("кол-во") or lc.get("шт")
                if qty_col:
                    ff_df[qty_col] = pd.to_numeric(ff_df[qty_col], errors="coerce").fillna(0.0).astype(float)
                    if seller_col in ff_df.columns:
                        ff_df[seller_col] = ff_df[seller_col].map(_norm_id)
                    if wb_col in ff_df.columns:
                        ff_df[wb_col] = ff_df[wb_col].map(_norm_id)
                    for _, r in ff_df.iterrows():
                        s_val = _norm_id(r.get(seller_col)) if seller_col else None
                        w_val = _norm_id(r.get(wb_col)) if wb_col else None
                        q_val = float(r.get(qty_col, 0.0) or 0.0)
                        key = (s_val or "", w_val or "")
                        ff_stock[key] = ff_stock.get(key, 0.0) + q_val
            if ff_stock:
                logs.append(f"Остатки ФФ: прочитано {len(ff_stock)} SKU")
            else:
                logs.append("Остатки ФФ: лист не найден, считаем остаток ФФ = 0")
        except Exception:
            pass

        df_base["Склад"] = df_base["Склад"].astype(str)

        results: Dict[str, pd.DataFrame] = {}

        for wh_name in (selected_wh if selected_wh else [None]):
            if wh_name is None:
                subset = df_base.copy()
            else:
                subset = df_base[df_base["Склад"] == wh_name].copy()
            if subset.empty:
                continue

            rec_rows: List[Dict[str, object]] = []
            for _, row in subset.iterrows():
                seller = str(row["Артикул продавца"]).strip()
                wb = str(row["Артикул WB"]).strip()
                key = (seller, wb)

                avg = float(row["Средние продажи в день"])
                stock_now = float(row["Остаток на сегодня"])
                coef = float(row.get("Коэф. склада", 0.0))

                minstock = minstock_map.get(key, 0.0)
                moq = moq_map.get(key, 0.0)
                freq = freq_map.get(wh_name or "", 0)
                horizon = acceptance_days + freq

                target = minstock + avg * horizon
                order_raw = target - stock_now
                order = max(0.0, order_raw)
                if moq > 0:
                    order = math.ceil(order / moq) * moq

                rec_rows.append(
                    {
                        "Артикул продавца": seller,
                        "Артикул WB": wb,
                        "Склад": wh_name if wh_name is not None else str(row["Склад"]),
                        "Средние продажи в день": avg,
                        "MinStock": minstock,
                        "Горизонт, дни": horizon,
                        "MOQ": moq,
                        "Коэф. склада": coef,
                        "Вес склада": 0.0,
                        "Остаток на сегодня": stock_now,
                        "Рекомендация, шт": int(order),
                        "Рекомендация с учётом ФФ": 0,
                    }
                )
                try:
                    last_row = rec_rows[-1]
                    s_key = _norm_id(last_row.get("Артикул продавца"))
                    w_key = _norm_id(last_row.get("Артикул WB"))
                    ff_key = (s_key or "", w_key or "")
                    last_row["Остаток ФФ"] = float(ff_stock.get(ff_key, 0.0))
                except Exception:
                    try:
                        rec_rows[-1]["Остаток ФФ"] = 0.0
                    except Exception:
                        pass

            sheet_key = wh_name or "Рекомендации"
            results[sheet_key] = pd.DataFrame(rec_rows)
            results[sheet_key]["Рекомендация, шт"] = (
                pd.to_numeric(results[sheet_key]["Рекомендация, шт"], errors="coerce")
                .fillna(0)
                .astype(int)
            )

            # Расчёт веса склада среди выбранных складов
            if "Коэф. склада" in results[sheet_key].columns:
                coeffs = results[sheet_key]["Коэф. склада"].astype(float).fillna(0)
                if len(results[sheet_key]) == 1:
                    results[sheet_key]["Вес склада"] = 1.0
                else:
                    total = coeffs.sum()
                    if total > 0:
                        results[sheet_key]["Вес склада"] = coeffs / total
                    else:
                        results[sheet_key]["Вес склада"] = 0.0
            else:
                results[sheet_key]["Вес склада"] = 0.0
            # ---- FF-логирование и расчёт FF-рекомендации ----
            # Для каждого SKU в этом листе сначала логируем FF и теоретический заказ,
            # затем считаем "Рекомендация с учётом ФФ" через _distribute_ff_for_sku.
            if "Рекомендация с учётом ФФ" in results[sheet_key].columns:
                results[sheet_key]["Рекомендация с учётом ФФ"] = \
                    results[sheet_key]["Рекомендация, шт"]

                if "Артикул продавца" in results[sheet_key].columns and "Артикул WB" in results[sheet_key].columns:
                    ff_stats = []
                    grouped = results[sheet_key].groupby(
                        ["Артикул продавца", "Артикул WB"],
                        dropna=False,
                        as_index=False,
                        sort=False,
                    )
                    for (sku_s, sku_w), group_df in grouped:
                        idx = group_df.index
                        try:
                            ff_val = float(group_df["Остаток ФФ"].iloc[0] or 0.0)
                        except Exception:
                            ff_val = 0.0

                        order_theory_sum = int(group_df["Рекомендация, шт"].sum())
                        ff_stats.append(
                            {
                                "seller": str(sku_s or "").strip(),
                                "wb": str(sku_w or "").strip(),
                                "ff": ff_val,
                                "demand": order_theory_sum,
                            }
                        )

                        if ff_val <= 0:
                            # Остатка на ФФ нет — рекомендация с учётом ФФ = 0
                            results[sheet_key].loc[idx, "Рекомендация с учётом ФФ"] = 0
                            continue

                        df_sku = results[sheet_key].loc[idx]
                        ff_series = _distribute_ff_for_sku(df_sku, ff_val)
                        results[sheet_key].loc[idx, "Рекомендация с учётом ФФ"] = ff_series

                    fully_covered = 0
                    partially_covered = 0
                    no_ff = 0
                    deficit = []
                    for s in ff_stats:
                        demand = s["demand"]
                        ff_val = s["ff"]
                        if demand <= 0:
                            continue
                        if ff_val <= 0:
                            no_ff += 1
                        elif ff_val >= demand:
                            fully_covered += 1
                        else:
                            partially_covered += 1
                            coverage = ff_val / demand if demand > 0 else 0.0
                            s["coverage"] = coverage
                            deficit.append(s)

                    logs.append(
                        f"FF-итог: полностью покрыто {fully_covered} SKU, частично покрыто {partially_covered}, без покрытия {no_ff}"
                    )

                    deficit_sorted = sorted(deficit, key=lambda s: s.get("coverage", 0.0))
                    top_deficit = deficit_sorted[:5]
                    for s in top_deficit:
                        logs.append(
                            f"FF-дефицит: {s['seller']}/{s['wb']} — FF={s['ff']}, спрос={s['demand']}, покрытие={s.get('coverage', 0.0):.0%}"
                        )
        if not results:
            results["Рекомендации"] = pd.DataFrame(
                columns=[
                    "Артикул продавца",
                    "Артикул WB",
                    "Склад",
                    "Средние продажи в день",
                    "MinStock",
                    "Горизонт, дни",
                    "MOQ",
                    "Коэф. склада",
                    "Вес склада",
                    "Остаток на сегодня",
                    "Остаток ФФ",
                    "Рекомендация, шт",
                    "Рекомендация с учётом ФФ",
                ]
            )

        # Формируем служебную сводную таблицу по всем листам рекомендаций
        ff_summary = pd.DataFrame()
        summary_rows = []
        for sheet_name, df_wh in results.items():
            if df_wh is None or df_wh.empty:
                continue
            tmp = df_wh.copy()
            # на всякий случай добавим название листа как метку склада-источника
            tmp["_Лист"] = sheet_name
            summary_rows.append(tmp)
        if summary_rows:
            ff_summary = pd.concat(summary_rows, ignore_index=True)

        # Подливаем в FF_Сводку общий Остаток ФФ по SKU (по Артикулу WB)
        if not ff_summary.empty and not ff_table.empty:
            try:
                ff_summary["Артикул WB"] = ff_summary["Артикул WB"].astype(str).str.strip()
                ff_table_wb = ff_table[["Артикул WB", "Остаток ФФ"]].copy()
                ff_table_wb["Артикул WB"] = ff_table_wb["Артикул WB"].astype(str).str.strip()
                ff_summary = ff_summary.merge(
                    ff_table_wb,
                    on="Артикул WB",
                    how="left",
                )
            except Exception:
                # если что-то пошло не так — просто не добавляем колонку
                pass

        # Сортируем FF_Сводку по SKU (Артикул WB, затем Артикул продавца)
        if not ff_summary.empty:
            try:
                ff_summary["Артикул WB"] = ff_summary["Артикул WB"].astype(str).str.strip()
                ff_summary["Артикул продавца"] = ff_summary["Артикул продавца"].astype(str).str.strip()
                ff_summary = ff_summary.sort_values(
                    ["Артикул WB", "Артикул продавца"]
                ).reset_index(drop=True)
            except Exception:
                pass

        # Таблица остатков ФФ по SKU для служебного листа «Остатки ФФ»
        # базовая таблица остатков ФФ
        ff_table = pd.DataFrame(
            [
                {
                    "Артикул продавца": k[0],
                    "Артикул WB": k[1],
                    "Количество": v,
                }
                for k, v in ff_stock.items()
            ]
        ) if ff_stock else pd.DataFrame(
            columns=["Артикул продавца", "Артикул WB", "Количество"]
        )

        # Явная колонка остатков ФФ для визуализации
        if not ff_table.empty:
            ff_table["Остаток ФФ"] = ff_table["Количество"]

        # ---- Добавляем колонку "Хватает на все" ----
        if not ff_table.empty:

            ff_table["Хватает на все"] = "Нет"
            for idx, row in ff_table.iterrows():
                wb = str(row["Артикул WB"]).strip()
                ff_qty = float(row["Количество"] or 0.0)

                # суммарная теоретическая рекомендация по выбранным складам
                total_demand = 0.0
                for df_wh in results.values():
                    try:
                        mask = (
                            df_wh["Артикул WB"].astype(str).str.strip() == wb
                        )
                        total_demand += float(df_wh.loc[mask, "Рекомендация, шт"].sum() or 0.0)
                    except Exception:
                        pass

                if ff_qty >= total_demand:
                    ff_table.at[idx, "Хватает на все"] = "Да"

            # ---- Добавляем колонки по выбранным складам с теоретическими рекомендациями ----
            if selected_wh:
                for wh_name in selected_wh:
                    col_name = str(wh_name)
                    if col_name not in ff_table.columns:
                        ff_table[col_name] = 0.0

                for idx, row in ff_table.iterrows():
                    wb = str(row["Артикул WB"]).strip()
                    for wh_name, df_wh in results.items():
                        if wh_name not in selected_wh:
                            continue
                        col_name = str(wh_name)
                        try:
                            mask = (
                                df_wh["Артикул WB"].astype(str).str.strip() == wb
                            )
                            demand_sum = float(df_wh.loc[mask, "Рекомендация, шт"].sum() or 0.0)
                            if demand_sum != 0.0:
                                ff_table.at[idx, col_name] = demand_sum
                        except Exception:
                            continue

        # Финальная сортировка таблицы по SKU ПЕРЕД записью листа
        try:
            ff_table["Артикул WB"] = ff_table["Артикул WB"].astype(str).str.strip()
            ff_table["Артикул продавца"] = ff_table["Артикул продавца"].astype(str).str.strip()
            ff_table = ff_table.sort_values([
                "Артикул WB", "Артикул продавца"
            ]).reset_index(drop=True)
        except Exception:
            pass

        out = BytesIO()
        try:
            writer = pd.ExcelWriter(out, engine="xlsxwriter")
        except Exception:
            writer = pd.ExcelWriter(out, engine="openpyxl")

        def _clean(name: str) -> str:
            raw = str(name)
            bad = ":*?/\\[]"
            cleaned = "".join(ch for ch in raw if ch not in bad).strip()
            return cleaned[:31] or "Склад"

        def _autofit_sheet(writer, sheet_name, df):
            if df is None or df.empty:
                return
            widths = []
            for col in df.columns:
                try:
                    maxlen = max(df[col].astype(str).map(len).max(), len(str(col)))
                except Exception:
                    maxlen = len(str(col))
                widths.append(min(maxlen + 2, 60))
            try:
                ws = writer.sheets.get(sheet_name)
                if ws is not None and hasattr(ws, "set_column"):
                    for idx, w in enumerate(widths):
                        ws.set_column(idx, idx, max(8, w))
                    return
            except Exception:
                pass
            try:
                from openpyxl.utils import get_column_letter
                ws = writer.book[sheet_name]
                for idx, w in enumerate(widths, start=1):
                    ws.column_dimensions[get_column_letter(idx)].width = max(8, w)
            except Exception:
                pass

        with writer:
            # основные листы по складам / сценариям
            for wh_name, df_wh in results.items():
                sheet_name = _clean(wh_name)
                # колонка «Остаток ФФ» нужна для расчётов, но в итоговом Excel не показываем
                df_to_write = df_wh.drop(columns=["Остаток ФФ"], errors="ignore")
                df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)
                _autofit_sheet(writer, sheet_name, df_to_write)

            # служебная сводка по FF и рекомендациям
            if not ff_summary.empty:
                # ограничим служебную вкладку только нужными колонками, если они есть
                cols = []
                for name in [
                    "Артикул продавца",
                    "Артикул WB",
                    "Склад",
                    "Коэф. склада",
                    "Вес склада",
                    "Остаток на сегодня",
                    "Рекомендация, шт",
                    "Рекомендация с учётом ФФ",
                    "Остаток ФФ",
                    "_Лист",
                ]:
                    if name in ff_summary.columns:
                        cols.append(name)
                ff_sheet = ff_summary[cols] if cols else ff_summary
                summary_name = _clean("FF_Сводка")
                ff_sheet.to_excel(writer, sheet_name=summary_name, index=False)
                _autofit_sheet(writer, summary_name, ff_sheet)

            # отдельный лист с остатками ФФ по SKU
            if not ff_table.empty:
                ff_raw_name = _clean("Остатки ФФ")
                ff_table.to_excel(writer, sheet_name=ff_raw_name, index=False)
                _autofit_sheet(writer, ff_raw_name, ff_table)
        out.seek(0)
        token = secrets.token_urlsafe(16)
        _memory_artifacts[token] = {
            "data": out.getvalue(),
            "filename": "WB_Replenishment_Recommendations.xlsx"
        }
        result_df = df_base.copy()
        logs.append(f"Рекомендации сформированы: {len(result_df)} строк")
        # FF отключён: вырезаем постобработку и удаляем столбец FF из результата
        # FF отключён: удаляем колонку тихо, без логов
        try:
            if "Рекомендация с учётом остатков FF" in result_df.columns:
                result_df.drop(columns=["Рекомендация с учётом остатков FF"], inplace=True, errors="ignore")
        except Exception:
            pass
        return {"ok": True, "log": "\n".join(logs), "download_token": token}
    except Exception as e:
        tb = traceback.format_exc()
        return JSONResponse({"ok": False, "log": "\n".join(logs + [f'Ошибка: {e}', 'TRACEBACK:', tb])}, status_code=500)
