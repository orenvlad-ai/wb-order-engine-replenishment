import io
import logging
import os
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from zipfile import BadZipFile


logger = logging.getLogger(__name__)


_DETAIL_REQUIRED = ["Склад", "Артикул продавца", "Артикул WB", "Заказали, шт"]
_DAILY_REQUIRED = ["Артикул продавца", "Артикул WB"]
_SNAPSHOT_REQUIRED = ["Артикул продавца", "Артикул WB", "Склад", "Остаток"]
_SNAPSHOT_ALIASES = {
    "Артикул продавца": {
        "артикул продавца",
        "артикул поставщика",
        "арт продавца",
    },
    "Артикул WB": {
        "артикул wb",
        "арт wb",
        "артикул wildberries",
        "артикул wb.",
    },
    "Склад": {
        "склад",
        "склады",
        "склад поставки",
    },
    "Остаток": {
        "остаток",
        "остатки",
        "остаток на складе",
        "доступный остаток",
        "количество",
        "в наличии",
    },
}
_INTRANSIT_COLUMNS = ["Артикул продавца", "Артикул WB", "Склад", "Количество"]
_INTRANSIT_ALIASES = {
    "Артикул продавца": {
        "артикул продавца",
        "артикул поставщика",
        "артикул",
        "арт",
    },
    "Артикул WB": {
        "артикул wb",
        "арт wb",
        "код номенклатуры",
        "код товара",
        "артикул wildberries",
        "артикул wb.",
    },
    "Количество": {
        "количество, шт.",
        "количество",
        "кол-во",
        "количество шт",
        "qty",
        "шт",
    },
}
_TARGET_SHEETS = {
    "детальная информация": ("Детальная информация", _DETAIL_REQUIRED),
    "остатки по дням": ("Остатки по дням", _DAILY_REQUIRED),
}


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")
    df.columns = [str(col).strip() for col in df.columns]
    return df


def _normalize_header_value(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\n", " ").replace("\r", " ")
    text = text.strip().strip(":")
    text = " ".join(text.lower().split())
    return text


def _match_aliases(row: Iterable[object], aliases: Dict[str, set[str]]) -> Optional[Dict[str, int]]:
    normalized = [_normalize_header_value(value) for value in row]
    mapping: Dict[str, int] = {}
    for index, value in enumerate(normalized):
        if not value:
            continue
        for target, options in aliases.items():
            if target in mapping:
                continue
            if value in options:
                mapping[target] = index
                break
    if len(mapping) == len(aliases):
        return mapping
    return None


def _build_snapshot_dataframe(rows: List[List[object]]) -> Optional[pd.DataFrame]:
    if not rows:
        return None

    for idx, row in enumerate(rows):
        mapping = _match_aliases(row, _SNAPSHOT_ALIASES)
        if not mapping:
            continue

        records: List[Dict[str, object]] = []
        for data_row in rows[idx + 1 :]:
            if not data_row or all(cell in (None, "") for cell in data_row):
                continue
            record: Dict[str, object] = {}
            empty = True
            for target in _SNAPSHOT_REQUIRED:
                position = mapping.get(target)
                value = data_row[position] if position is not None and position < len(data_row) else None
                if isinstance(value, str):
                    value = value.strip()
                if value not in (None, ""):
                    empty = False
                record[target] = value
            if empty:
                continue
            records.append(record)

        if not records:
            return pd.DataFrame(columns=_SNAPSHOT_REQUIRED)
        return pd.DataFrame(records)

    return None


def _load_snapshot_excel(file_bytes: bytes) -> Optional[pd.DataFrame]:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return None

    try:
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            df = _build_snapshot_dataframe(rows)
            if df is not None:
                return df
    finally:
        workbook.close()
    return None


def _load_snapshot_csv(file_bytes: bytes) -> Optional[pd.DataFrame]:
    csv_df = _try_read_csv(file_bytes)
    if csv_df is None:
        return None

    if csv_df.empty:
        rows = [list(csv_df.columns)]
    else:
        normalized_df = csv_df.where(~csv_df.isna(), None)
        rows = [list(csv_df.columns)] + normalized_df.astype(object).values.tolist()
    return _build_snapshot_dataframe(rows)


def _row_contains_all_required(row: pd.Series, required: list[str]) -> bool:
    values = {str(v).strip() for v in row if pd.notna(v)}
    return all(req in values for req in required)


def _ensure_header(df: pd.DataFrame, required: list[str]) -> pd.DataFrame:
    if df.empty:
        return df

    columns = [str(col).strip() for col in df.columns]
    if all(req in columns for req in required):
        df.columns = columns
        return df

    first_row = df.iloc[0]
    if _row_contains_all_required(first_row, required):
        df = df.iloc[1:].copy()
        df.columns = [str(v).strip() for v in first_row]
        return df

    df.columns = columns
    return df


def _match_target_sheet(sheet_name: str) -> tuple[str, list[str]] | None:
    key = sheet_name.strip().lower()
    return _TARGET_SHEETS.get(key)


def _load_excel(file_bytes: bytes) -> Dict[str, pd.DataFrame]:
    excel = pd.ExcelFile(io.BytesIO(file_bytes))
    sheets: Dict[str, pd.DataFrame] = {}
    for original_name in excel.sheet_names:
        df = excel.parse(original_name)
        df = _clean_dataframe(df)
        target = _match_target_sheet(original_name)
        if target:
            normalized_name, required = target
            df = _ensure_header(df, required)
            sheets[normalized_name] = df
    return sheets


def _try_read_csv(file_bytes: bytes) -> pd.DataFrame | None:
    encodings = [None, "utf-8", "utf-16", "cp1251"]
    delimiters = [";", ",", "\t"]
    for encoding in encodings:
        for delimiter in delimiters:
            try:
                buffer = io.BytesIO(file_bytes)
                df = pd.read_csv(buffer, encoding=encoding, sep=delimiter)
                if df.empty:
                    continue
                return _clean_dataframe(df)
            except Exception:
                continue
    return None


def read_stock_history(file_bytes: bytes, filename: str | None = None) -> Dict[str, pd.DataFrame]:
    """Прочитать отчёт «История остатков» и вернуть релевантные листы."""

    if not file_bytes:
        return {}

    try:
        sheets = _load_excel(file_bytes)
        if sheets:
            return sheets
    except Exception:
        sheets = {}

    csv_df = _try_read_csv(file_bytes)
    if csv_df is None:
        return {}

    csv_df = _ensure_header(csv_df, _DETAIL_REQUIRED)
    return {"Детальная информация": csv_df}


def read_stock_snapshot(file_bytes: bytes, filename: str | None = None) -> Optional[pd.DataFrame]:
    """Прочитать отчёт «Остатки по складам» и вернуть подготовленный DataFrame."""

    if not file_bytes:
        return None

    df = _load_snapshot_excel(file_bytes)
    if df is None:
        df = _load_snapshot_csv(file_bytes)

    if df is None:
        return None

    df = _clean_dataframe(df)
    missing = [column for column in _SNAPSHOT_REQUIRED if column not in df.columns]
    if missing:
        return None

    return df[_SNAPSHOT_REQUIRED]


def read_intransit_file(data: bytes, filename: str) -> pd.DataFrame:
    """Прочитать файл с поставками в пути и вернуть подготовленный DataFrame."""

    def _warehouse_name() -> str:
        base = os.path.splitext(os.path.basename(filename or ""))[0]
        return base.strip()

    def _prepare_dataframe(df: pd.DataFrame | None) -> pd.DataFrame | None:
        if df is None or df.empty:
            return None

        cleaned = df.dropna(how="all").dropna(axis=1, how="all")
        if cleaned.empty:
            return None

        cleaned.columns = [str(col).strip() for col in cleaned.columns]
        normalized = {_normalize_header_value(col): col for col in cleaned.columns}

        mapping: Dict[str, str] = {}
        for target, aliases in _INTRANSIT_ALIASES.items():
            for alias in aliases:
                column = normalized.get(alias)
                if column:
                    mapping[target] = column
                    break

        if "Количество" not in mapping:
            return None
        if "Артикул продавца" not in mapping and "Артикул WB" not in mapping:
            return None

        result = pd.DataFrame(index=cleaned.index)

        def _normalize_article(series: pd.Series) -> pd.Series:
            def _convert(value: object) -> Optional[str]:
                if pd.isna(value):
                    return None
                text = str(value).strip()
                if not text:
                    return None
                if text.lower() in {"nan", "none", "null"}:
                    return None
                return text or None

            return series.map(_convert)

        def _normalize_quantity(series: pd.Series) -> pd.Series:
            prepared = (
                series.astype(str)
                .str.replace("\xa0", "", regex=False)
                .str.replace(" ", "", regex=False)
                .str.replace(",", ".", regex=False)
                .str.strip()
            )
            numeric = pd.to_numeric(prepared, errors="coerce").fillna(0)
            return numeric.round().astype(int)

        seller_column = mapping.get("Артикул продавца")
        wb_column = mapping.get("Артикул WB")
        qty_column = mapping["Количество"]

        if seller_column:
            result["Артикул продавца"] = _normalize_article(cleaned[seller_column])
        else:
            result["Артикул продавца"] = pd.Series([None] * len(cleaned), index=cleaned.index)

        if wb_column:
            result["Артикул WB"] = _normalize_article(cleaned[wb_column])
        else:
            result["Артикул WB"] = pd.Series([None] * len(cleaned), index=cleaned.index)

        result["Количество"] = _normalize_quantity(cleaned[qty_column])
        result["Склад"] = _warehouse_name() or ""

        mask = (
            result["Количество"] > 0
        ) & (~result["Артикул продавца"].isna() | ~result["Артикул WB"].isna())

        filtered = result.loc[mask, _INTRANSIT_COLUMNS]
        filtered.attrs["intransit"] = True
        return filtered.reset_index(drop=True)

    # Пытаемся прочитать как Excel (шапка может быть на первой или второй строке)
    for header in (0, 1):
        try:
            excel_df = pd.read_excel(io.BytesIO(data), header=header, engine="openpyxl")
        except (BadZipFile, ValueError, TypeError):
            excel_df = None
        except Exception:
            continue
        prepared = _prepare_dataframe(excel_df)
        if prepared is not None:
            return prepared

    # Если Excel не подошёл, пробуем CSV с разными кодировками
    for header in (0, 1):
        for encoding in ("utf-8-sig", "utf-8", "cp1251", "utf-16"):
            try:
                csv_df = pd.read_csv(
                    io.BytesIO(data),
                    header=header,
                    encoding=encoding,
                    sep=None,
                    engine="python",
                )
            except Exception:
                continue
            prepared = _prepare_dataframe(csv_df)
            if prepared is not None:
                return prepared

    return pd.DataFrame(columns=_INTRANSIT_COLUMNS)


def read_fulfillment_stock_file(data: bytes, filename: str) -> pd.DataFrame:
    """
    Парсит файл остатков Фулфилмента.
    Ожидаемые колонки (любой из синонимов):
      - Артикул продавца: ["артикул продавца", "артикул поставщика", "артикул"]
      - Артикул WB (опционально): ["артикул wb", "артикул wb.", "артикул вб", "код товара"]
      - Количество:       ["количество", "остаток", "кол-во", "шт"]
    Возвращает DataFrame с колонками ["Артикул продавца", "Артикул WB", "Количество"].
    """

    df = pd.DataFrame()

    # Пробуем прочитать как Excel с заголовками на первой или второй строке
    for header in (0, 1):
        try:
            df = pd.read_excel(io.BytesIO(data), header=header, engine="openpyxl")
            if not df.empty:
                break
        except (BadZipFile, ValueError):
            df = pd.DataFrame()
        except Exception:
            continue

    # Если Excel не прочитан, пробуем CSV с популярными кодировками
    if df.empty:
        for encoding in ("utf-16", "utf-8-sig", "utf-8", "cp1251"):
            try:
                df = pd.read_csv(
                    io.BytesIO(data),
                    header=0,
                    encoding=encoding,
                    sep=None,
                    engine="python",
                )
                if not df.empty:
                    break
            except Exception:
                continue

    if df.empty:
        return pd.DataFrame(columns=["Артикул продавца", "Артикул WB", "Количество"])

    normalized = {str(column).strip().lower(): column for column in df.columns}

    def _pick_column(options: Iterable[str]) -> Optional[str]:
        for option in options:
            if option in normalized:
                return normalized[option]
        return None

    seller_column = _pick_column(["артикул продавца", "артикул поставщика", "артикул"])
    wb_column = _pick_column(["артикул wb", "артикул wb.", "артикул вб", "код товара"])
    quantity_column = _pick_column(["количество", "остаток", "кол-во", "шт"])

    if seller_column is None or quantity_column is None:
        return pd.DataFrame(columns=["Артикул продавца", "Артикул WB", "Количество"])

    result = pd.DataFrame()
    result["Артикул продавца"] = df[seller_column]
    result["Артикул WB"] = df[wb_column] if wb_column in df.columns else None
    result["Количество"] = (
        pd.to_numeric(
            df[quantity_column]
            .astype(str)
            .str.replace("\xa0", "", regex=False)
            .str.replace(" ", "", regex=False)
            .str.replace(",", ".", regex=False),
            errors="coerce",
        )
        .fillna(0)
        .astype(int)
    )

    mask = (~result["Артикул продавца"].isna()) & (result["Количество"] > 0)
    return result.loc[mask, ["Артикул продавца", "Артикул WB", "Количество"]].reset_index(drop=True)


# Удалено: ридер справочника SKU больше не используется.
